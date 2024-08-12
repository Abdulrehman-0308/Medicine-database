from flask import Flask, request, jsonify
from flask_restplus import Api, Resource
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

app = Flask(_name_)
api = Api(app)

class Config:
    SQLALCHEMY_DATABASE_URI = 'postgresql://username:password@localhost/dbname'
    SQLALCHEMY_TRACK_MODIFICATIONS = False

app.config.from_object(Config)

db = SQLAlchemy(app)

class Medication(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_name = db.Column(db.String(255), nullable=False)
    name = db.Column(db.String(255), nullable=False)
    type = db.Column(db.String(255), nullable=False)
    quantity = db.Column(db.String(255), nullable=False)
    dosage = db.Column(db.String(255), nullable=False)
    frequency = db.Column(db.String(255), nullable=False)
    duration = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    updated_at = db.Column(db.DateTime, default=db.func.current_timestamp(), onupdate=db.func.current_timestamp())

@api.route('/upload')
class Upload(Resource):
    def post(self):
        file = request.files['file']
        filename = secure_filename(file.filename)
        wb = load_workbook(filename=file)
        sheet = wb.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            product_name, dosage, frequency, duration = row
            name, type, quantity = product_name.split(' ', 2)
            data.append({
                'product_name': product_name,
                'name': name,
                'type': type,
                'quantity': quantity,
                'dosage': dosage,
                'frequency': frequency,
                'duration': duration
            })
        db.session.bulk_save_objects([Medication(**d) for d in data])
        db.session.commit()
        return jsonify({'message': 'Data uploaded successfully'})

if _name_ == '_main_':
    app.run(debug=True)
